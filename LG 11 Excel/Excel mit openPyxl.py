"""
OpenPyxl ist eine Python-Bibliothek, die zum Lesen und Schreiben von Excel-Dateien im xlsx-Format verwendet wird. 
Hier sind einige grundlegende Beispiele, wie man OpenPyxl verwendet:

Installation
Bevor du OpenPyxl verwenden kannst, musst du es installieren. Du kannst das mit pip tun:


pip install openpyxl

"""

import openpyxl

# Schreiben in eine Excel-Datei

# Erstelle ein neues Arbeitsbuch und wähle das aktive Arbeitsblatt aus
workbook = openpyxl.Workbook()
sheet = workbook.active

# Schreibe Daten in eine Zelle
sheet['A1'] = 'Hallo, Welt!'

# Speichere die Datei
workbook.save('example.xlsx')


# Lesen einer Excel-Datei

# Lade die Excel-Datei
workbook = openpyxl.load_workbook('example.xlsx')

# Wähle das aktive Arbeitsblatt aus
sheet = workbook.active

# Lese den Wert einer bestimmten Zelle
cell_value = sheet['A1'].value
print(cell_value)

# Durchlaufe alle Zeilen und Spalten
for row in sheet.iter_rows(values_only=True):
    print(row)


# Arbeiten mit mehreren Arbeitsblättern

# Erstelle ein neues Arbeitsbuch
workbook = openpyxl.Workbook()

# Erstelle ein neues Arbeitsblatt
workbook.create_sheet(title='Neues Blatt')

# Wähle das neue Arbeitsblatt aus
sheet = workbook['Neues Blatt']

# Schreibe Daten in das neue Arbeitsblatt
sheet['A1'] = 'Daten im neuen Blatt'

# Speichere die Datei
workbook.save('example.xlsx')

# Arbeiten mit Zellbereichen

# Lade die Excel-Datei
workbook = openpyxl.load_workbook('example.xlsx')
sheet = workbook.active

# Lese einen Bereich von Zellen
cell_range = sheet['A1:B2']
for row in cell_range:
    for cell in row:
        print(cell.value)

# Schreibe in einen Bereich von Zellen
for row in sheet.iter_rows(min_row=1, max_row=2, min_col=1, max_col=2):
    for cell in row:
        cell.value = 'Neuer Wert'

# Speichere die Datei
workbook.save('example.xlsx')
