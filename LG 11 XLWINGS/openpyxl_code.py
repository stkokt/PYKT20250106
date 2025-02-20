#from openpyxl import Workbook

#wb = Workbook()

# grab the active worksheet
#ws = wb.active

# Data can be assigned directly to cells
#ws['A1'] = 42



import openpyxl

wb = openpyxl.open('sample.xlsx')
ws = wb.active
#print(list(ws.values))


# Rows can also be appended
#ws.append([4, 5, 6])
ws.insert_rows(4, 5, 6)

wb.save("sample.xlsx")
# Python types will automatically be converted
# import datetime
# ws['A2'] = datetime.datetime.now()

# Save the file

